from openpyxl_boardgames.utils import clean_data, normalize
from openpyxl_boardgames.config import URL1, URL2, URL3, COLUMNS, OBJECT_FLOAT, FILL_NA, NEW_TYPE
from openpyxl_boardgames.excel import create_ws, scatter_chart, bar_chart, data_validation, radar_chart
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl import load_workbook
import time
import pandas as pd



def main():
    try :
        t = time.perf_counter()

        df_main = pd.read_csv(URL1, sep=';')
        df_sec = pd.read_csv(URL2, sep=',')
        df_thr = pd.read_csv(URL3, sep=',')
        print(f"Données récupérées en {time.perf_counter()-t:.2f}s")

        df_main=df_main.merge(df_sec,how="left",left_on='ID',right_on='game_id')
        df_main=df_main.merge(df_thr,how="left",left_on='ID',right_on='id')
        print("Jointure effectuée")

        df_main=df_main[COLUMNS]

        df_main=clean_data(df_main, OBJECT_FLOAT, FILL_NA, NEW_TYPE)
        print("Données nettoyées")

        df_main=normalize(df_main,"Rating Average","Rating Average n")
        df_main=normalize(df_main,"Complexity Average","Complexity Average n")
        df_main=normalize(df_main,"Users Rated", "Users Rated n")
        df_main=normalize(df_main,"Owned Users", "Owned Users n")
        df_main=normalize(df_main,"Min Players", "Min Players n")
        df_main=normalize(df_main,"Max Players", "Max Players n")
        df_main=normalize(df_main,"Play Time", "Play Time n")
        df_main=normalize(df_main,"Min Age", "Min Age n")
        print("Données normalisées")

        pd.DataFrame.to_excel(df_main, excel_writer='template.xlsx', sheet_name='DATA', index=False)
        print("Données exportées au format .xlsx")


        wb=load_workbook("template.xlsx")
        print("Workbook chargé")

        create_ws(wb,"TDB","output/test/test.xlsx")
        print("Feuille TDB créée")

        ws_tdb=wb["TDB"]
        ws_tdb["E1"]="Physionomie des jeux de sociétés"
        ws_tdb.merge_cells(range_string="E1:M3")
        ws_tdb["E1"].font = Font(
            name='Calibri',
            size=28,
            bold=True,
            italic=True
        )
        ws_tdb["E1"].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        ws_tdb["B5"]="Filtre famille"
        ws_tdb.merge_cells(range_string="B5:C7")
        ws_tdb["B5"].font = Font(
            name='Calibri',
            size=18,
            bold=True
        )
        ws_tdb["B5"].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        ws_tdb["E5"]="Nombre de jeux"
        ws_tdb.merge_cells(range_string="E5:F7")
        ws_tdb["E5"].font = Font(
            name='Calibri',
            size=18,
            bold=True
        )
        ws_tdb["E5"].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)


        ws_tdb["H5"]="Proportion"
        ws_tdb.merge_cells(range_string="H5:I7")
        ws_tdb["H5"].font = Font(
            name='Calibri',
            size=18,
            bold=True
        )
        ws_tdb["H5"].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        ws_tdb["K5"]="Rang BGG médian"
        ws_tdb.merge_cells(range_string="K5:L7")
        ws_tdb["K5"].font = Font(
            name='Calibri',
            size=18,
            bold=True
        )
        ws_tdb["K5"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        ws_tdb["N5"]="Année moyenne de parution"
        ws_tdb.merge_cells(range_string="N5:O7")
        ws_tdb["N5"].font = Font(
            name='Calibri',
            size=14,
            bold=True
        )
        ws_tdb["N5"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        ws_tdb.merge_cells(range_string="B8:C9")
        ws_tdb.merge_cells(range_string="E8:F9")
        ws_tdb.merge_cells(range_string="H8:I9")
        ws_tdb.merge_cells(range_string="K8:L9")
        ws_tdb.merge_cells(range_string="N8:O9")
        print("Mise en forme des KPIs")

        data_validation(wb=wb,worksheet_data=wb["DATA"],worksheet_tdb=wb["TDB"],col_data=14,col_tdb=26,data_title="Liste_domaine",where_tdb="B8",save_as="output/test/test.xlsx")

        jeux = dict(df_main["Domains"].value_counts())
        longueur_max=max(jeux.values())
        formula = '=_xlfn.FILTER(DATA!A:AH,DATA!N:N=TDB!B8,"")'
        ws_tdb["AA1"] = ArrayFormula(f"AA1:BH{longueur_max}", formula)
        print("Ajout des données de DATA dans TDB pour être filtrées")

        scatter_chart(wb,wb["TDB"],wb["TDB"],where="A12",col_x=37,col_y=35,min_row=1,max_row=1000,save_as="output/test/test.xlsx")
        print("Scatter plot tracé")

        bar_chart(wb,wb["TDB"],wb["TDB"],where="J12",col_x=28,col_y=38,min_row=1,max_row=10,save_as="output/test/test.xlsx")
        print("Bar plot tracé")

        ws_tdb["E8"]=f'={longueur_max}-COUNTIF(AG:AG,"<>"&NA())'
        ws_tdb["E8"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        ws_tdb["H8"]="=ROUND(E8/23327,2)*100"
        ws_tdb["H8"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        ws_tdb["K8"]='=_xlfn.AGGREGATE(12,6,AJ:AJ)'
        ws_tdb["K8"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        ws_tdb["N8"]='=_xlfn.AGGREGATE(12,6,AC:AC)'
        ws_tdb["N8"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        print("KPI ajoutés")

        ws_tdb["X1"]="Note moyenne normalisée"
        ws_tdb["Y1"]='=_xlfn.AGGREGATE(1,6,BA:BA)'
        ws_tdb["X2"]="Complexité moyenne normalisée"
        ws_tdb["Y2"]='=_xlfn.AGGREGATE(1,6,BB:BB)'
        ws_tdb["X3"]="Nombre moyen minimal de joueurs normalisé"
        ws_tdb["Y3"]='=_xlfn.AGGREGATE(1,6,BE:BE)'
        ws_tdb["X4"]="Nombre moyen maximal de joueurs normalisé"
        ws_tdb["Y4"]='=_xlfn.AGGREGATE(1,6,BF:BF)'
        ws_tdb["X5"]="Temps de partie moyen normalisé"
        ws_tdb["Y5"]='=_xlfn.AGGREGATE(1,6,BG:BG)'
        ws_tdb["X6"]="Age minimal moyen normalisé"
        ws_tdb["Y6"]='=_xlfn.AGGREGATE(1,6,BH:BH)'
        wb.save("output/test/test.xlsx")

        radar_chart(wb,wb["TDB"],wb["TDB"],where="A37",min_row=1,max_row=5,min_col=24,save_as="output/test/test.xlsx")
        print("Radar chart tracé")

    except Exception as e:
        print(f"Erreur : {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
