from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import (
    ScatterChart,
    BubbleChart,
    Reference,
    #existe bien
    Series
)
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

def load_wb(path: str) -> Workbook:
    #ajouter la docstring
    return load_workbook(filename = path)


def create_ws(wb: Workbook, sheet_name: str, save_as: str):
    # ajouter la docstring
    ws = wb.create_sheet(sheet_name, 0)
    wb.save(save_as)

def copy_columns(wb: Workbook, 
                 ws_data: Worksheet, 
                 ws_destination: Worksheet, 
                 col_data: int, 
                 col_destination: int, 
                 min_row: int, 
                 save_as: str) -> None :
    # ajouter la docstring
    row_line = min_row
    for row in ws_data.iter_rows(min_row=min_row, max_row=ws_data.max_row):
        val=row[col_data-1].value
        #pas une vraie erreur ci-dessous
        ws_destination.cell(row=row_line, column=col_destination).value=val
        row_line+=1
    wb.save(save_as)



def scatter_chart(
        wb: Workbook, 
        worksheet_chart: Worksheet, 
        worksheet_data: Worksheet, 
        where: str, 
        col_x: int, 
        col_y: int, 
        min_row: int, 
        max_row: int, 
        save_as: str) -> None :
    
    # ajouter la docstring
    # Pas convaincu par le résultat, je ne pense pas que ça fasse réellement un nuage de points, ça relie les points entre eux.
    c1 = ScatterChart()
    # c1.title = "Complexité vs Note utilisateur"
    c1.style = 5
    c1.legend = None
    c1.y_axis.title = 'Note utilisateur'
    c1.x_axis.title = 'Complexité'
    c1.width = 20
    c1.height = 15

    x_values = Reference(worksheet_data, min_col=col_x, min_row=min_row, max_row=max_row)
    y_values = Reference(worksheet_data, min_col=col_y, min_row=min_row, max_row=max_row)
    series = Series(y_values, x_values, title_from_data=True)

    series.marker.symbol = "circle" 
    series.marker.size = 5
    series.graphicalProperties.line.noFill = True
    
    c1.series.append(series)

    worksheet_chart.add_chart(c1, where)

    wb.save(save_as)


def bubble_chart(wb: Workbook, 
        worksheet_chart: Worksheet, 
        worksheet_data: Worksheet, 
        where: str, 
        col_x: int, 
        col_y: int,
        col_size: int, 
        min_row: int, 
        max_row: int, 
        save_as: str) -> None :
    # ajouter la docstring
    c1 = BubbleChart()
    # c1.title = "Complexité vs Note utilisateur"
    c1.style = 1
    c1.legend = None
    c1.y_axis.title = 'Note utilisateur'
    c1.x_axis.title = 'Complexité'
    c1.width = 20
    c1.height = 15

    x_values = Reference(worksheet_data, min_col=col_x, min_row=min_row, max_row=max_row)
    y_values = Reference(worksheet_data, min_col=col_y, min_row=min_row, max_row=max_row)
    size = Reference(worksheet_data,min_col=col_size, min_row=min_row, max_row=max_row)
    series = Series(y_values, x_values, size, title_from_data=True)
    c1.series.append(series)

    worksheet_chart.add_chart(c1, where)

    wb.save(save_as)


def add_filter(wb: Workbook, worksheet: Worksheet, range: str, save_as: str) -> None:
    # ajouter la docstring
    filters = worksheet.auto_filter
    # peu importe la ligne de fin mais il faut les bonnes colonnes
    filters.ref = range
    wb.save(save_as)