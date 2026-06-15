"""Création d'un filtre sous la forme d'une validation de données"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation


def data_validation(ws: Worksheet, worksheet_data: Worksheet,where: str, col_data: int, col_ws: int)->None:
    """_summary_

    Args:
        ws (Worksheet): Worksheet où mettre le filtre
        worksheet_data (Worksheet): Worksheet où aller chercher les données
        where (str): Emplacement du filtre
        col_data (int): Colonne des données
        col_ws (int): Colonne où sont copiées les données de worksheet_data
    """
    domains=[]
    seen=set()

    for row in range(2, worksheet_data.max_row + 1):
        value = worksheet_data.cell(row=row, column=col_data).value
        if value is not None and str(value).strip() != "":
            value=str(value)
            if value not in seen :
                seen.add(value)
                domains.append(value)
    
    domains.sort()

    ws.cell(row = 1, column = col_ws).value = "Famille"
    for i, j in enumerate(domains, start= 2):
        ws.cell(row = i, column = col_ws).value = j
    
    dv = DataValidation(type="list", formula1="TDB!$Z$2:$Z$41", allow_blank=False)
    dv.prompt = "Choisissez une famille"
    dv.promptTitle = "Famille"
    dv.error = "Valeur non autorisée"
    dv.errorTitle = "Erreur"

    ws.add_data_validation(dv)
    dv.add(where)
