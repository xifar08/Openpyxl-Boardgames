from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import (
    ScatterChart,
    BubbleChart,
    Reference,
    BarChart,
    #existe bien
    Series
)
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter



def load_wb(path: str) -> Workbook:
    #ajouter la docstring
    # c'est un wrapper, pas nécessaire
    return load_workbook(filename = path)


def create_ws(wb: Workbook, sheet_name: str, save_as: str):
    # ajouter la docstring
    # c'est un wrapper, pas nécessaire
    ws = wb.create_sheet(sheet_name, 0)
    wb.save(save_as)


def copy_columns(
        wb: Workbook, 
        ws_data: Worksheet, 
        ws_destination: Worksheet, 
        col_data: int, 
        col_destination: int, 
        min_row: int, 
        save_as: str) -> None :
    """_summary_

    Args:
        wb (Workbook): _description_
        ws_data (Worksheet): _description_
        ws_destination (Worksheet): _description_
        col_data (int): _description_
        col_destination (int): _description_
        min_row (int): _description_
        save_as (str): _description_
    """
    row_line = min_row
    for row in ws_data.iter_rows(min_row=min_row, max_row=ws_data.max_row):
        val=row[col_data-1].value
        #pas une vraie erreur ci-dessous
        ws_destination.cell(row=row_line, column=col_destination).value=val
        row_line+=1
    wb.save(save_as)



def scatter_chart(
        wb: Workbook, 
        worksheet_chart: Worksheet, 
        worksheet_data: Worksheet, 
        where: str, 
        col_x: int, 
        col_y: int, 
        min_row: int, 
        max_row: int, 
        save_as: str) -> None :
    
    
    # Pas convaincu par le résultat, je ne pense pas que ça fasse réellement un nuage de points, ça relie les points entre eux.
    c1 = ScatterChart()
    # c1.title = "Complexité vs Note utilisateur"
    #c1.style = 5
    c1.legend = None
    c1.y_axis.title = 'Note utilisateur'
    c1.x_axis.title = 'Complexité'
    c1.width = 20
    c1.height = 15

    x_values = Reference(worksheet_data, min_col=col_x, min_row=min_row, max_row=max_row)
    y_values = Reference(worksheet_data, min_col=col_y, min_row=min_row, max_row=max_row)
    series = Series(y_values, x_values, title_from_data=False)

    series.marker.symbol = "circle" 
    series.marker.size = 5
    series.graphicalProperties.line.noFill = True
    
    c1.series.append(series)

    worksheet_chart.add_chart(c1, where)

    wb.save(save_as)


def bubble_chart(
        wb: Workbook, 
        worksheet_chart: Worksheet, 
        worksheet_data: Worksheet, 
        where: str, 
        col_x: int, 
        col_y: int,
        col_size: int, 
        min_row: int, 
        max_row: int, 
        save_as: str) -> None :
    """_summary_

    Args:
        wb (Workbook): _description_
        worksheet_chart (Worksheet): _description_
        worksheet_data (Worksheet): _description_
        where (str): _description_
        col_x (int): _description_
        col_y (int): _description_
        col_size (int): _description_
        min_row (int): _description_
        max_row (int): _description_
        save_as (str): _description_
    """
    c1 = BubbleChart()
    # c1.title = "Complexité vs Note utilisateur"
    c1.style = 1
    c1.legend = None
    c1.y_axis.title = 'Note utilisateur'
    c1.x_axis.title = 'Complexité'
    c1.width = 20
    c1.height = 15

    x_values = Reference(worksheet_data, min_col=col_x, min_row=min_row, max_row=max_row)
    y_values = Reference(worksheet_data, min_col=col_y, min_row=min_row, max_row=max_row)
    size = Reference(worksheet_data,min_col=col_size, min_row=min_row, max_row=max_row)
    series = Series(y_values, x_values, size, title_from_data=True)
    c1.series.append(series)

    worksheet_chart.add_chart(c1, where)

    wb.save(save_as)


def add_filter(wb: Workbook, worksheet: Worksheet, range: str, save_as: str) -> None:
    """_summary_

    Args:
        wb (Workbook): _description_
        worksheet (Worksheet): _description_
        range (str): _description_
        save_as (str): _description_
    """
    filters = worksheet.auto_filter
    # peu importe la ligne de fin mais il faut les bonnes colonnes
    filters.ref = range
    wb.save(save_as)


def data_validation(
        wb: Workbook, 
        worksheet_data: Worksheet, 
        worksheet_tdb: Worksheet, 
        col_data: int,
        col_tdb: int,
        data_title : str,
        where_tdb :str,
        save_as: str) -> None:
    """_summary_

    Args:
        wb (Workbook): _description_
        worksheet_data (Worksheet): _description_
        worksheet_tdb (Worksheet): _description_
        col_data (int): _description_
        save_as (str): _description_
    """
    data=[]
    seen=set()

    for row in range(2, worksheet_data.max_row + 1):
        value = worksheet_data.cell(row=row, column=col_data).value
        if value is not None and str(value).strip() != "":
            value=str(value)
            if value not in seen :
                seen.add(value)
                data.append(value)
    
    data.sort()

    worksheet_tdb.cell(row = 1, column = col_tdb).value = data_title
    for i, j in enumerate(data, start= 2):
        worksheet_tdb.cell(row = i, column = col_tdb).value = j

    if data_title in wb.defined_names:
        del wb.defined_names[data_title]

    end_row = 2 + len(data) - 1
    col_letter = get_column_letter(col_tdb)
    ref_data = "'" + worksheet_tdb.title + "'!$" + col_letter + "$2:$" + col_letter + "$" + str(end_row)
    wb.defined_names[data_title] = DefinedName(data_title, attr_text=ref_data)
    
    dv = DataValidation(type="list", formula1=f"={data_title}", allow_blank=True)
    dv.prompt = "Choisis un domaine"
    dv.promptTitle = "Domain"
    dv.error = "Valeur non autorisée"
    dv.errorTitle = "Erreur"

    worksheet_tdb.add_data_validation(dv)
    dv.add(where_tdb)
    wb.save(save_as)


def bar_chart(
        wb: Workbook, 
        worksheet_chart: Worksheet, 
        worksheet_data: Worksheet, 
        where: str, 
        col_x: int, 
        col_y: int, 
        min_row: int, 
        max_row: int, 
        save_as: str) -> None :
    
    
    c1 = BarChart()
    c1.type="col"
    c1.legend = None
    c1.y_axis.title = 'Nombre de jeux possédés'
    c1.x_axis.title = 'Jeux'
    c1.width = 20
    c1.height = 15

    x_values = Reference(worksheet_data, min_col=col_x, min_row=min_row, max_row=max_row)
    y_values = Reference(worksheet_data, min_col=col_y, min_row=min_row, max_row=max_row)

    c1.add_data(y_values, titles_from_data=False)
    c1.set_categories(x_values)

    worksheet_chart.add_chart(c1, where)

    wb.save(save_as)
