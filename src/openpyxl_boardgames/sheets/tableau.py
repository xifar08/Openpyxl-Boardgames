"""Construction de la feuille TABLEAU : ajout des titres et des formules"""

from openpyxl_boardgames.config import (
    SHEET_TABLEAU,
    TITLE_TABLEAU,
    TITLE_CELLS_TABLEAU,
    KPI_CELLS_TABLEAU,
    KPI_FORMULA_TABLEAU,
    KPI_FORMULA_RANGE
    )

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.worksheet.formula import ArrayFormula

def build_tdb_tableau(wb: Workbook) -> None:
    """Orchéstrtaionde la contruction de l'onglet TABLEAU
    Création de la feuille dans le workbook, ajout des titres des colonnes puis des formules.

    Args:
        wb (Workbook): Workbook sans feuille TABLEAU
    """

    ws_tab=wb.create_sheet(SHEET_TABLEAU,1)

    add_titles(ws=ws_tab)

    add_kpis(ws=ws_tab)


def add_titles(ws: Worksheet)-> None :
    """Ajout des titres des colonnes

    Args:
        ws (Worksheet): Onglet TABLEAU
    """
    for title, col in zip(TITLE_TABLEAU, TITLE_CELLS_TABLEAU):
        ws[col]=title
        ws[col].font = Font(
            name='Calibri',
            size=11,
            bold=True
        )


def add_kpis(ws: Worksheet)-> None :
    """Ajout des formules dans les colonnes

    Args:
        ws (Worksheet): Onglet TABLEAU
    """
    for formula, col, rg in zip(KPI_FORMULA_TABLEAU, KPI_CELLS_TABLEAU, KPI_FORMULA_RANGE):
        ws[col]=ArrayFormula(rg, formula)