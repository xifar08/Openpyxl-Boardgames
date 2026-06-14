from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl_boardgames.config import (
    SHEET_TDB,
    SHEET_DATA,
    TITLE_TDB,
    TITLE_CELLS_TDB,
    TITLE_MERGE_TDB,
    KPI_FORMULA_TDB,
    KPI_CELLS_TDB,
    KPI_MERGE_TDB,
    FORMULA_FILTER,
    RADCHART_CELLS,
    RADCHART_TITLES,
    RADCHART_FORMULA1,
    RADCHART_FORMULA2
    )
from openpyxl_boardgames.components.data_validation import data_validation
from openpyxl_boardgames.components.charts import bar_chart, scatter_chart, radar_chart


def build_tdb_sheet(wb: Workbook) -> None:
    """_summary_

    Args:
        wb (Workbook): _description_
    """

    ws_tdb=wb.create_sheet(SHEET_TDB,0)

    ws_tdb.sheet_view.showGridLines = False

    add_titles(ws_tdb)

    add_kpis(ws_tdb)

    add_radchart_data(ws_tdb)

    data_validation(ws=ws_tdb, worksheet_data=wb[SHEET_DATA],where="B8",col_data=14, col_ws=26)

    ws_tdb["AA1"] = ArrayFormula("AA1", FORMULA_FILTER)

    scatter_chart(ws_tdb,ws_tdb,where="A12",col_x=37,col_y=35,min_row=1,max_row=1000)

    bar_chart(ws_tdb,ws_tdb,where="J12",col_x=28,col_y=38,min_row=1,max_row=10)


    radar_chart(ws_tdb,ws_tdb,where="F37",min_row=1,max_row=7,min_col=23)



def add_titles(ws: Worksheet)-> None:
    """_summary_

    Args:
        ws (Worksheet): _description_
    """
    ws["A1"]="Pour faire fonctionner le TdB, veuillez double cliquer en AA1 et appuyer sur entrer avant de choisir un filtre."
    ws.merge_cells(range_string="A1:C4")
    ws["A1"].alignment = Alignment(horizontal='left', vertical='center',wrapText=True)

    ws["E1"]="Physionomie des jeux de sociétés"
    ws.merge_cells(range_string="E1:M3")
    ws["E1"].font = Font(
        name='Calibri',
        size=28,
        bold=True,
        italic=True
    )
    ws["E1"].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

    for title, col, merge in zip(TITLE_TDB, TITLE_CELLS_TDB, TITLE_MERGE_TDB):
        ws[col]=title
        ws.merge_cells(range_string=merge)
        ws[col].font = Font(
            name='Calibri',
            size=14,
            bold=True
        )
        ws[col].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)


def add_kpis(ws: Worksheet)-> None:
    """_summary_

    Args:
        ws (Worksheet): _description_
    """
    ws.merge_cells(range_string="B8:C9")
    ws["B8"].font = Font(
        name='Calibri',
        size=9,
        color='00FF9900',
        bold=True
    )
    ws["B8"].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)


    for formula, col, merge in zip(KPI_FORMULA_TDB, KPI_CELLS_TDB, KPI_MERGE_TDB):
        ws[col]=formula
        ws.merge_cells(range_string=merge)
        ws[col].font = Font(
            name='Calibri',
            size=20,
            bold=True
        )
        ws[col].alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
    
    ws["H8"].number_format = '0.0%'
    ws["N8"].number_format = '0'

def add_radchart_data(ws: Worksheet)-> None:
    """_summary_

    Args:
        ws (Worksheet): _description_
    """
    ws["X1"]="Profil de la famille"
    ws["Y1"]="Profil du jeu le mieux classé de la famille"

    for title,row,formula1, formula2 in zip(RADCHART_TITLES,RADCHART_CELLS,RADCHART_FORMULA1,RADCHART_FORMULA2):
        ws["W"+row]=title
        ws["X"+row]=formula1
        ws["Y"+row]=formula2





# ws_tdb["W2"]="Note moyenne normalisée"
# ws_tdb["X2"]='=_xlfn.AGGREGATE(1,6,BA:BA)'
# ws_tdb["Y2"]='=TABLEAU!L2'
# ws_tdb["W3"]="Complexité moyenne normalisée"
# ws_tdb["X3"]='=_xlfn.AGGREGATE(1,6,BB:BB)'
# ws_tdb["Y3"]='=TABLEAU!M2'
# ws_tdb["W4"]="Nombre moyen minimal de joueurs normalisé"
# ws_tdb["X4"]='=_xlfn.AGGREGATE(1,6,BE:BE)'
# ws_tdb["Y4"]='=TABLEAU!N2'
# ws_tdb["W5"]="Nombre moyen maximal de joueurs normalisé"
# ws_tdb["X5"]='=_xlfn.AGGREGATE(1,6,BF:BF)'
# ws_tdb["Y5"]='=TABLEAU!O2'
# ws_tdb["W6"]="Temps de partie moyen normalisé"
# ws_tdb["X6"]='=_xlfn.AGGREGATE(1,6,BG:BG)'
# ws_tdb["Y6"]='=TABLEAU!P2'
# ws_tdb["W7"]="Age minimal moyen normalisé"
# ws_tdb["X7"]='=_xlfn.AGGREGATE(1,6,BH:BH)'
# ws_tdb["Y7"]='=TABLEAU!Q2'
# wb.save("output/test/test.xlsx")